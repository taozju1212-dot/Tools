"""
Excel 导出 / 导入模块

Sheet 结构（单 Sheet，所有模式依次排列）：
  Row: ##MODE | mode_id | beat_time | step   ← 模式标记行（橙色）
  Row: 模块 | 动作编号 | 动作名称 | 0.0 | 0.5 | ... | beat_time | 起始时间 | 前级动作 | 动作时间
  Row: 数据行 × N
  Row: （空行分隔）
"""
from __future__ import annotations
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import AppData, ModeConfig, TimelineRow

_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="FF1565C0")  # dark blue
_FILL_MODE    = PatternFill(fill_type="solid", fgColor="FFFF8F00")  # amber
_FILL_METAHDR = PatternFill(fill_type="solid", fgColor="FFE3F2FD")  # light blue
_FONT_HEADER  = Font(bold=True, color="FFFFFFFF", size=9)
_FONT_MODE    = Font(bold=True, color="FFFFFFFF", size=10)
_FONT_META    = Font(bold=True, color="FF1565C0", size=8)
_FONT_DATA    = Font(size=9)

MODE_MARKER = "##MODE"  # sentinel in col A of mode header rows


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor="FF" + hex_color.lstrip("#").upper())


def _time_cols_for(beat_time: int, step: float) -> list[float]:
    times, t = [], 0.0
    while t <= beat_time:
        times.append(round(t, 1))
        t = round(t + step, 1)
    return times


def _fmt_t(t: float) -> str:
    return str(int(t)) if t == int(t) else str(t)


# ── Export ─────────────────────────────────────────────────────────────────────

def export_to_excel(
    sections_data: list[tuple[ModeConfig, list[list[Optional[str]]], list[float]]],
    app_data: AppData,
    path: str,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "节拍时序"

    cur_row = 1

    for sec_idx, (cfg, color_grid, time_cols) in enumerate(sections_data):
        mode_obj = next((m for m in app_data.modes if m.id == cfg.mode_id), None)
        mode_name = mode_obj.name if mode_obj else ""

        meta_cols = 3  # 起始时间, 前级动作, 动作时间
        total_cols = 3 + len(time_cols) + meta_cols  # A=模块 B=动作编号 C=动作名称 ...

        # ── mode marker row ────────────────────────────────────────────────────
        ws.cell(cur_row, 1, MODE_MARKER)
        ws.cell(cur_row, 2, cfg.mode_id)
        ws.cell(cur_row, 3, cfg.beat_time)
        ws.cell(cur_row, 4, cfg.step)
        ws.cell(cur_row, 5, f"模式: {cfg.mode_id} {mode_name}  节拍:{cfg.beat_time}S  步长:{cfg.step}S")
        # merge display cell across remaining columns for readability
        ws.merge_cells(start_row=cur_row, start_column=5,
                       end_row=cur_row, end_column=total_cols)
        for ci in range(1, total_cols + 1):
            c = ws.cell(cur_row, ci)
            c.fill = _FILL_MODE
            c.font = _FONT_MODE
            c.alignment = _CENTER
            c.border = _BORDER
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # ── column header row ──────────────────────────────────────────────────
        headers = (
            ["模块", "动作编号", "动作名称"]
            + [_fmt_t(t) for t in time_cols]
            + ["起始时间", "前级动作", "动作时间"]
        )
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(cur_row, ci, h)
            if ci <= 3:
                c.fill = _FILL_HEADER
                c.font = _FONT_HEADER
            else:
                is_meta = ci > 3 + len(time_cols)
                c.fill = _FILL_METAHDR if is_meta else _FILL_HEADER
                c.font = _FONT_META if is_meta else _FONT_HEADER
            c.alignment = _CENTER
            c.border = _BORDER
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

        # ── data rows ──────────────────────────────────────────────────────────
        for row_idx, tr in enumerate(cfg.rows):
            mod_obj  = next((m for m in app_data.modules if m.id == tr.module_id), None)
            mod_disp = f"{tr.module_id} {mod_obj.name}" if mod_obj else tr.module_id
            act_obj  = app_data.get_action(tr.action_key)
            act_name = act_obj.name if act_obj else ""

            # fixed cols
            for ci, val in enumerate([mod_disp, tr.action_key, act_name], start=1):
                c = ws.cell(cur_row, ci, val)
                c.font = _FONT_DATA
                c.alignment = _CENTER
                c.border = _BORDER

            # time cols with color
            row_colors = color_grid[row_idx] if row_idx < len(color_grid) else []
            for ti, hex_color in enumerate(row_colors):
                c = ws.cell(cur_row, 4 + ti)
                c.border = _BORDER
                c.alignment = _CENTER
                if hex_color:
                    c.fill = _hex_fill(hex_color)

            # meta cols
            meta_start_ci = 4 + len(time_cols)
            ws.cell(cur_row, meta_start_ci,     tr.start_time if tr.start_time is not None else "").border = _BORDER
            ws.cell(cur_row, meta_start_ci + 1, tr.prev_action_key or "").border = _BORDER
            ws.cell(cur_row, meta_start_ci + 2, tr.duration).border = _BORDER
            for ci in range(meta_start_ci, meta_start_ci + 3):
                ws.cell(cur_row, ci).alignment = _CENTER
                ws.cell(cur_row, ci).font = _FONT_DATA

            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # blank separator row between modes
        cur_row += 1

    # ── column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    # Time columns + meta — determine max time cols across sections
    if sections_data:
        max_time = max(len(tc) for _, _, tc in sections_data)
        for ci in range(4, 4 + max_time):
            ws.column_dimensions[get_column_letter(ci)].width = 4
        # meta cols after the widest time section
        for offset, w in enumerate([9, 10, 8], start=1):
            ws.column_dimensions[get_column_letter(4 + max_time + offset - 1)].width = w

    ws.freeze_panes = "D2"
    wb.save(path)


# ── Import ─────────────────────────────────────────────────────────────────────

def import_from_excel(path: str) -> list[ModeConfig]:
    """Read an Excel file produced by export_to_excel and return ModeConfig list."""
    wb = load_workbook(path, data_only=True)
    # Find the main sheet (first sheet named "节拍时序" or just first sheet)
    ws = wb["节拍时序"] if "节拍时序" in wb.sheetnames else wb.active

    configs: list[ModeConfig] = []
    rows = list(ws.iter_rows(values_only=True))

    i = 0
    while i < len(rows):
        row = rows[i]
        if row and str(row[0]).strip() == MODE_MARKER:
            # Parse mode marker: col B=mode_id, col C=beat_time, col D=step
            mode_id   = str(row[1]).strip() if row[1] is not None else ""
            beat_time = int(row[2]) if row[2] is not None else 30
            step      = float(row[3]) if row[3] is not None else 0.5

            time_cols = _time_cols_for(beat_time, step)
            meta_start = 3 + len(time_cols)  # 0-indexed: cols 0,1,2 = 模块/动作编号/动作名称

            cfg = ModeConfig(mode_id=mode_id, beat_time=beat_time, step=step)
            i += 2  # skip marker row + header row

            while i < len(rows):
                dr = rows[i]
                # blank row = section end
                if dr is None or all(v is None or str(v).strip() == "" for v in dr):
                    break
                # next mode marker = section end
                if dr[0] is not None and str(dr[0]).strip() == MODE_MARKER:
                    break

                # col 0: module display like "A 移液器" → extract module_id = first char
                mod_raw = str(dr[0]).strip() if dr[0] is not None else ""
                module_id = mod_raw[0] if mod_raw else ""

                # col 1: action key like "A00"
                action_key = str(dr[1]).strip() if dr[1] is not None else ""
                action_no  = action_key[1:] if len(action_key) >= 3 else ""

                # meta cols (0-indexed)
                def _val(idx):
                    return dr[idx] if idx < len(dr) else None

                raw_start = _val(meta_start)
                start_time: Optional[float] = None
                if raw_start is not None and str(raw_start).strip() not in ("", "None"):
                    try:
                        start_time = float(raw_start)
                    except (ValueError, TypeError):
                        pass

                raw_prev = _val(meta_start + 1)
                prev_key: Optional[str] = None
                if raw_prev is not None and str(raw_prev).strip() not in ("", "None"):
                    prev_key = str(raw_prev).strip()

                raw_dur = _val(meta_start + 2)
                duration = 1.0
                if raw_dur is not None:
                    try:
                        duration = float(raw_dur)
                    except (ValueError, TypeError):
                        pass

                cfg.rows.append(TimelineRow(
                    module_id=module_id,
                    action_no=action_no,
                    start_time=start_time,
                    prev_action_key=prev_key,
                    duration=duration,
                ))
                i += 1

            configs.append(cfg)
        else:
            i += 1

    return configs
